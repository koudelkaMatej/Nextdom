from asyncore import write
from cmath import nan
import datetime
from hashlib import new
import os.path
import os
import copy
from time import time
from unittest import skip, skipUnless
import pandas as pd
import numpy as np
from statistics import mean, median, stdev
# https://docs.scipy.org/doc/scipy/reference/generated/scipy.stats.norm.html     https://docs.scipy.org/doc/scipy/reference/generated/scipy.stats.poisson.html
from scipy.stats import poisson, norm
import scipy
import numpy as np
import openpyxl
import re

from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule,FormulaRule
from openpyxl.utils.dataframe import dataframe_to_rows

# mluvit o statistice a proc jsem ji pouzil
# implementace + testování ověření
#user acceptance testing (rozhovor s Pavlou) # nacitani csv -> UTF8 - problem proto encoding


def mutation(AAchange):
    if AAchange != '':
        CRM = pd.read_csv("static/defaultCSV/ClinicalRelevantMutation.txt")
        digitsPattern = r'\d+'
        lettersPattern = r'[A-Za-z]+'
        AA2 = re.findall(digitsPattern, AAchange)
        letter = re.findall(lettersPattern, AAchange)
        if "p" in letter:
            letter.remove("p")
        letter = ''.join(letter)
        AA = []

    # Loop over the characters in the string
        for c in letter:
            # If the character is not already in the unique list, add it
            if c not in AA:
                AA.append(c)

    # Convert the list to a numpy array
        AA = np.array(AA)

        AA1 = AA[0]
        AA3 = AA[-1]
        AA2 = AA2[0]

        pat = ''.join([str(AA1), ''.join(AA2), str(AA3)])
        for col in CRM.columns:
            if pat in CRM[col].values:
                mut = 'CRM'
            else:
                mut = ''
        return pat, mut
    pat = ''
    mut = ''
    return pat, mut

def findAK(posL, posH):
    NLAK = cellvalue(posL, 2, False)
    TLAK = cellvalue(posL, 4, False)
    NHAK = cellvalue(posH, 2, True)
    THAK = cellvalue(posH, 4, True)
    try:    
        PosLAK = TLAK + str(int(float(NLAK)))
        PosHAK = THAK + str(int(float(NHAK)))
    except:
        PosLAK = nan
        PosHAK=  nan
    return PosLAK, PosHAK

def pos(X, H):
    resMod = X % 3
    if resMod == 1:
        posun = 0
    elif resMod == 0:
        posun = 1
    else:
        posun = 2
    if H and resMod != 1:
        posun -= 6
    elif H:
        posun -= 3
    return posun

def cellvalue(X, col, H):
    posAA = pd.read_csv('static/defaultCSV/PoziceAA.csv', sep="\t")
    trans = pos(X, H) + X - 271
    val = str(posAA.iloc[trans-1, col-1])
    return val



## začatek mainu


matrice = pd.read_csv("static/defaultCSV/matrice.csv", sep='\t', header=None,engine='python')
CRM = pd.read_csv("static/defaultCSV/ClinicalRelevantMutation.txt")
array = np.array(CRM)
#scipy.io.savemat('C:/Users/koude/OneDrive - České vysoké učení technické v Praze/FBMI_MAG/Diplomova_prace/Klempond/git/Nextdom/HTMLaCSSaFlask/AnalyzeMatlab/Nextdom2.0/CRM.mat', {'df_array': array})
zac = 1
rozsah = []
meze_popis = []
LoD = []
U = []
V = []
W_all = []


def report_part1_save_xlsx(partsPath,nameOfRun,logger):
    os.makedirs(partsPath+"/Results", exist_ok=True)
    soubor = partsPath+ '/Results/' + nameOfRun +  '_P1'+".xlsx"
    part1Path = partsPath+"/Part1"
    files = [os.path.join(part1Path, x) for x in os.listdir(part1Path) if x.endswith(".txt")]  # ziskam cely obsah slozky na txt soubory s celou cestou
    filenames = [x for x in os.listdir(part1Path) if x.endswith(".txt")]
    T_sablona = pd.read_excel("static/defaultCSV/sablona.xlsx")
    try:
        for i in range(len(filenames)):
            logger.info('File loaded: ' + str(i+1)+'/'+str(len(filenames)) + '    ' +filenames[i] )
            cesta_in = part1Path+ "/"+filenames[i]
            with open(cesta_in, 'r') as f:
                data = ''
                for j in range(3):
                    data += f.readline()

            textak =  pd.read_csv(cesta_in, delimiter='\t',skiprows=3)
            T = T_sablona
            T =pd.concat([T,  textak.drop(columns=textak.columns[-1])], axis=1)
            T.insert(loc=len(T.columns), column='', value=(''))
            T['A#(F)'] = T['A#(F;R)'].str.extract(r'^([^;]*)', expand=False).astype(float)
            T['A#(R)'] = T['A#(F;R)'].str.extract(r';(.*)', expand=False).astype(float)
            #T.drop(columns=['A#(F;R)'], inplace=True)

            # extract values from "C#(F;R)" column
            T['C#(F)'] = T['C#(F;R)'].str.extract(r'^([^;]*)', expand=False).astype(float)
            T['C#(R)'] = T['C#(F;R)'].str.extract(r';(.*)', expand=False).astype(float)
            #T.drop(columns=['C#(F;R)'], inplace=True)

            # extract values from "G#(F;R)" column
            T['G#(F)'] = T['G#(F;R)'].str.extract(r'^([^;]*)', expand=False).astype(float)
            T['G#(R)'] = T['G#(F;R)'].str.extract(r';(.*)', expand=False).astype(float)
            #T.drop(columns=['G#(F;R)'], inplace=True)

            # extract values from "T#(F;R)" column
            T['T#(F)'] = T['T#(F;R)'].str.extract(r'^([^;]*)', expand=False).astype(float)
            T['T#(R)'] = T['T#(F;R)'].str.extract(r';(.*)', expand=False).astype(float)
            #T.drop(columns=['T#(F;R)'], inplace=True)

            T['--']= ''
            T['>']= ''
            T['A'] = T['A#(F)'] + T['A#(R)']
            T['C'] = T['C#(F)'] + T['C#(R)']
            T['G'] = T['G#(F)'] + T['G#(R)']
            T['T'] = T['T#(F)'] + T['T#(R)']
            T.insert(loc=len(T.columns), column=' ', value=(''))
            T['suma'] = T['A'] + T['C'] + T['G'] + T['T']
            T['//']= ''
            T['F'] = T['A#(F)'] + T['C#(F)'] + T['G#(F)'] + T['T#(F)']
            T['R'] = T['A#(R)'] + T['C#(R)'] + T['G#(R)'] + T['T#(R)']
            T.drop(columns=['T#(F;R)'], inplace=True)
            T.drop(columns=['G#(F;R)'], inplace=True)
            T.drop(columns=['C#(F;R)'], inplace=True)
            T.drop(columns=['A#(F;R)'], inplace=True)

            data = data.replace('\t', '\n')
            data = data.split("\n")
            hlav = np.array([[data[0], data[1]], [data[2], data[3]], [data[4], data[5]]])
            list_c = os.path.splitext(data[3])[0].split("_")[0]
            prumer=['Average', T['suma'].mean()]
            kde = (T[['F', 'R']].min(axis=1) >= 1000)  #pd.Series(kde).value_counts()[True]
            misto = T['Index'] * kde
            #mistoUpravene = [0, misto,0]
            zpos = np.where(np.concatenate(([0], misto, [0]))==[0] ) 
            grpidx = np.argmax(np.diff(zpos))
            y = misto[zpos[0][grpidx]:zpos[0][grpidx+1]-2]

            if y.empty:
                omezeni = [['Dolní mez:', np.nan, 0], ['Horní mez:', np.nan, 0]]
                AKlimits = [0, 0]
            else:
                low = T['Pos'][y.idxmin()]
                high = T['Pos'][y.idxmax()]+1
                try:
                    LAK, HAK = findAK(low, high)
                except ValueError as e:
                    raise Exception(e)
                except:
                    LAK = nan
                    HAK = nan
                omezeni = [['Dolní mez:', np.nan, low], ['Horní mez:', np.nan, high]]
                AKlimits = [LAK, HAK]

                        # create a new Excel file or append to an existing file
            if os.path.isfile(soubor):
                # create a new worksheet
                wb = openpyxl.load_workbook(soubor)
                if list_c in wb.sheetnames:
                    continue
                else:
                    worksheet  = wb.create_sheet(title=list_c) 
                for r_idx, row in enumerate(dataframe_to_rows(T, index=False, header=True), start=5):
                    for c_idx, value in enumerate(row, start=1):
                        worksheet.cell(row=r_idx, column=c_idx, value=value)

                # write header row to worksheet
                worksheet.cell(row=1, column=1, value=hlav[0][0])
                worksheet.cell(row=1, column=2, value=hlav[0][1])
                worksheet.cell(row=2, column=1, value=hlav[1][0])
                worksheet.cell(row=2, column=2, value=hlav[1][1])
                worksheet.cell(row=3, column=1, value=hlav[2][0])
                worksheet.cell(row=3, column=2, value=hlav[2][1])

                worksheet.cell(row=1, column=8, value=prumer[0])
                worksheet.cell(row=2, column=8, value=int(prumer[1]))

                worksheet.cell(row=1, column=11, value=omezeni[0][0])
                worksheet.cell(row=2, column=11, value=omezeni[1][0])
                worksheet.cell(row=1, column=13, value=omezeni[0][2])
                worksheet.cell(row=2, column=13, value=omezeni[1][2])

                worksheet.cell(row=1, column=14, value=AKlimits[0])
                worksheet.cell(row=2, column=14, value=AKlimits[1])
            else:
                wb = openpyxl.Workbook()
                if "Sheet" in wb.sheetnames:
                    wb.active.title = list_c
                worksheet = wb[list_c]

                for r_idx, row in enumerate(dataframe_to_rows(T, index=False, header=True), start=5):
                    for c_idx, value in enumerate(row, start=1):
                        worksheet.cell(row=r_idx, column=c_idx, value=value)
    # write header row to worksheet
                worksheet.cell(row=1, column=1, value=hlav[0][0])
                worksheet.cell(row=1, column=2, value=hlav[0][1])
                worksheet.cell(row=2, column=1, value=hlav[1][0])
                worksheet.cell(row=2, column=2, value=hlav[1][1])
                worksheet.cell(row=3, column=1, value=hlav[2][0])
                worksheet.cell(row=3, column=2, value=hlav[2][1])

                worksheet.cell(row=1, column=8, value=prumer[0])
                worksheet.cell(row=2, column=8, value=int(prumer[1]))

                worksheet.cell(row=1, column=11, value=omezeni[0][0])
                worksheet.cell(row=2, column=11, value=omezeni[1][0])
                worksheet.cell(row=1, column=13, value=omezeni[0][2])
                worksheet.cell(row=2, column=13, value=omezeni[1][2])
                
                worksheet.cell(row=1, column=14, value=AKlimits[0])
                worksheet.cell(row=2, column=14, value=AKlimits[1])
            
            wb.save(soubor)
            wb.close()
    except Exception as e:
        logger.error(e)
        error_message = "Pravděpodobně jste zadali špatný soubor/y.\nZkuste to znovu"
        raise Exception(error_message)

def report_part2_save_xlsx(partsPath, nameOfRun,limits,logger):
    
    if limits=="0":
        limits_typ='LoD'
    else:
        limits_typ='LoQ'
    part2Path = partsPath+"/Part2"
    cesta_out = partsPath+ '/Results'
    soubor = cesta_out+"/"+nameOfRun+"_P2_" +limits_typ +".xlsx"
    files = [os.path.join(part2Path, x) for x in os.listdir(part2Path) if x.endswith(".txt")]  # ziskam cely obsah slozky na txt soubory s celou cestou
    filenames = [x for x in os.listdir(part2Path) if x.endswith(".txt")]
    try:
        for i in range(len(filenames)):
            logger.info('File loaded: ' + str(i+1)+'/'+str(len(filenames)) + '    ' +filenames[i] )
            cesta_in = part2Path+ "/"+filenames[i]
            with open(cesta_in, 'r') as f:
                data = ''
                for j in range(3):
                    data += f.readline()
            data = data.replace('\t', '\n')
            data = data.split("\n")

            T = pd.read_csv(os.path.join(cesta_in), sep='\t', header=3)   
            if T.empty:
                aa = pd.DataFrame([np.nan]*len(T.columns)).T
                aa.columns = T.columns
                T = pd.concat([T, aa])
            T["Alt%"] = T["Alt%"].str.replace(',', '.').astype(float)
            T["Overall Score"] = T["Overall Score"].str.replace(',', '.').astype(float)
            spl = filenames[i].split('_')
            NG = ["Software:", f"NextDOM 2.0 {limits_typ} Application"]
            Project = ["Project:", f"{spl[0]}_{spl[1]}_{limits_typ}"]
            Cas = ["Datetime:", str(datetime.datetime.today().strftime('%d-%m-%Y %H:%M:%S'))]
            Tcell = T.values.tolist()
            sizeT = len(Tcell)
            names = list(T.columns)
            rozsah.append('A6:R' + str(sizeT+5))
            hlav = [NG, Project, Cas]
            list_c = spl[0]
            if os.path.exists(soubor):
        # Load the workbook
                workbook = openpyxl.load_workbook(soubor)
            else:
                # Create a new workbook
                workbook = openpyxl.Workbook()
                if "Sheet" in workbook.sheetnames:
                    workbook.active.title = spl[0]

            # Loop over unique values in the first column of T
                # Check if the sheet already exists
            if spl[0] in workbook.sheetnames:
                sheet = workbook[spl[0]]
            else:
                # Create a new sheet
                sheet = workbook.create_sheet(title=spl[0]) ##  workbook.create_sheet(title=spl[0]+"_"+spl[1]) 

            # Write data to the sheet
            sheet.cell(row=1, column=1, value=NG[0])
            sheet.cell(row=1, column=2, value=NG[1])
            sheet.cell(row=2, column=1, value=Project[0])
            sheet.cell(row=2, column=2, value=Project[1])
            sheet.cell(row=3, column=1, value=Cas[0])
            sheet.cell(row=3, column=2, value=Cas[1])

            # Write column headers to the sheet
            for col_num, col_name in enumerate(T.columns):
                sheet.cell(row=5, column=col_num+1, value=col_name)

        # Write data to the sheet
            for row_num, row_data in enumerate(T.values):
                for col_num, cell_value in enumerate(row_data):
                    sheet.cell(row=row_num+6, column=col_num+1, value=cell_value)

            # Save the workbook
            workbook.save(soubor)
        workbook.close()    # save the workbook  
    except Exception as e:
        logger.error(e)
        error_message = "Pravděpodobně jste zadali špatný soubor/y.\nZkuste to znovu"
        raise Exception(error_message)


def report_part3_save_xlsx(partsPath, nameOfRun,limits,logger):
    if limits=="0":
        koef=1
        limits_typ="LoD"
    else:
        koef=5
        limits_typ="LoQ"
    cesta_out = cesta_out = partsPath+ '/Results'
    nazev3 = os.path.join(cesta_out, nameOfRun + '_P1.xlsx')
    vystup = os.path.join(cesta_out, nameOfRun + '_P2_' + limits_typ + '.xlsx')
    xl = pd.ExcelFile(nazev3)
    sheet_names = xl.sheet_names

# Get the number of sheets in the Excel file
    pocet_listu = len(xl.sheet_names)
    xl.close()
    kteri = range(0, pocet_listu)
    try:
        for i in kteri:
            df = pd.read_excel(nazev3, sheet_name=sheet_names[i], header=None, skiprows=5, usecols='R:U')
            data_array = df.values
            U1 = pd.DataFrame(data_array, columns=['A', 'C', 'G', 'T'])
            df = pd.read_excel(nazev3, sheet_name=sheet_names[i], header=None, skiprows=5, usecols='H:O')
            data_array = df.values
            V1 = pd.DataFrame(data_array, columns=["A#(F)",	"A#(R)"	,"C#(F)",	"C#(R)"	,"G#(F)",	"G#(R)"	,"T#(F)",	"T#(R)"])
            df = pd.read_excel(nazev3, sheet_name=sheet_names[i], header=None, nrows=2,usecols='H:N')
            data_array = df.values
            bb= pd.DataFrame(data_array)
            meze_popis.append(bb)
            U.append(U1)
            V.append(V1)

        gen = pd.read_excel(nazev3,sheet_name=sheet_names[i], usecols='C', skiprows=4).values
        gen = pd.DataFrame(gen, columns=['ref. NK'])
        pozice = pd.read_excel(nazev3, sheet_name=sheet_names[i], usecols='F', skiprows=4).values
        pozice = pd.DataFrame(pozice, columns=['Pos'])
        idx = pd.read_excel(nazev3, sheet_name=sheet_names[i], usecols='E', skiprows=4).values
        idx = pd.DataFrame(idx, columns=['Index'])
        T = pd.read_excel("static/defaultCSV/seznam_genu.xlsx").values
        T = pd.DataFrame(T, columns=['Pos',"ref_NK","hodnoty"])
        kde = np.where(matrice.iloc[0, 0] == T['Pos'])[0][-1]
        kde2 = np.where(matrice.iloc[-1, 0] == T['Pos'])[0][-1] + 1
        
        ext_matrice_zac = pd.DataFrame({'col1': [], 'col2': [], 'col3': [], 'col4': []})
        ext_matrice_kon = pd.DataFrame({'col1': [], 'col2': [], 'col3': [], 'col4': []})
        for i in range(4):
            LoD.append([matrice.loc[matrice[1]==i+1, 2].mode()[0], matrice.loc[matrice[1]==i+1, 3].mode()[0]])
            ext_matrice_zac.loc[ext_matrice_zac['col3']==i+1, 'col3'] = LoD[i][0]
            ext_matrice_zac.loc[ext_matrice_zac['col4']==i+1, 'col4'] = LoD[i][1]
            ext_matrice_kon.loc[ext_matrice_kon['col3']==i+1, 'col3'] = LoD[i][0]
            ext_matrice_kon.loc[ext_matrice_kon['col4']==i+1, 'col4'] = LoD[i][1]

        if not ext_matrice_zac.empty and not ext_matrice_zac.isna().values.any() and not ext_matrice_kon.empty and not ext_matrice_kon.isna().values.any():
            matrice_all = pd.concat([ext_matrice_zac, matrice, ext_matrice_kon])
        else:
            matrice_all = matrice
        
        Rz = [U[i] for i in kteri]
        RzV = [V[i] for i in kteri]
        seznam = ['A', 'C', 'G', 'T']
        geny = np.zeros(len(gen))
        for g in range(len(seznam)):
            for i in range(0, len(gen)):
                if gen.iloc[i,0] == seznam[g]:
                    geny[i] = g+1
        forward=1000
        backward=1000
        matrice_uni = []
        matrice_uni_LoQ = []
        for i in range(len(Rz)):
            df = pd.read_excel(vystup, sheet_name=sheet_names[i])

            start_row = 5
            end_row = df.shape[0]
            AAchange = df.iloc[start_row-1:end_row, 17].values.tolist()
            s = len(AAchange)
            AAchange = ["" if x is np.NaN else x for x in AAchange]
            # Process data
            if s == 1:
                pat, mut = mutation(AAchange)
            else:
                pat, mut = zip(*map(mutation, AAchange))    
            normala = RzV[i]
            Forw = normala.iloc[:, 0::2]
            Forw = Forw.sum(axis=1)
            Backw = normala.iloc[:, 1::2]
            Backw = Backw.sum(axis=1)
            min_pocet = np.min([Forw, Backw], axis=0)
            kde = min_pocet >= 1000
            misto = idx.mul(kde, axis=0)    #mistoUpravene = [0, misto,0]  
            misto.loc[-1] = [0] * len(misto.columns)
            misto.index = misto.index + 1  # shift the existing rows down
            misto = misto.sort_index()
            misto.loc[misto.index.max() + 1] = [0] 
            zpos = np.where(misto==[0])
            grpidx = np.argmax(np.diff(zpos))
            y = misto[zpos[0][grpidx]:zpos[0][grpidx+1]-1]

            if not y.empty:  
                dolni_limit = np.min(y)
                horni_limit = np.max(y)
                W= [Rz[i].iloc[:,0], Rz[i].iloc[:,1],    Rz[i].iloc[:,2],   Rz[i].iloc[:,3],  pozice, pd.DataFrame(geny)]  
                W = pd.concat(W, axis=1)        
                #spatne = np.concatenate(range(0:dolni_limit-1), horni_limit+1:len(geny)])
                spatne = np.concatenate([np.arange(1, dolni_limit-1), np.arange(horni_limit+1, geny.shape[0])])
        
                # remove the rows in W corresponding to indices in spatne
                W = W.drop(index=spatne)
                # calculate the coefficient
                koeficient = np.sum(W.iloc[:, :4], axis=1) / 3000
                # create the matrix typ_zmen
                typ_zmen = np.array([[0, 1, 2, 3], [2, 3, 0, 1], [1, 0, 1, 0], [3, 2, 3, 2]])
                poc_tranzice = np.zeros((len(W ), len(Rz)))
                poc_transverze = np.zeros((len(W ), len(Rz)))
                for j in range(1, 5):
                    kde = W.iloc[:, 5] == j  # A, C, G, T
                    poc_tranzice[kde, i] = W[kde].iloc[:,typ_zmen[1, j-1]]
                    poc_transverze[kde, i] = W[kde].iloc[:,typ_zmen[2, j-1]] + W[kde].iloc[:,typ_zmen[3, j-1]]
                poc_tranzice_norm = poc_tranzice[:, i] / koeficient
                poc_transverze_norm = poc_transverze[:, i] / koeficient
                W = pd.concat([W, pd.DataFrame({'poc_tranzice_norm': poc_tranzice_norm, 'poc_transverze_norm': poc_transverze_norm})], axis=1)
                W_all.append(W)
                kde = (matrice_all.iloc[:, 0] < W.iloc[0, 4]) | (matrice_all.iloc[:, 0] > W.iloc[-1, 4])
                kde = ~kde
                matrice_uni.append(matrice_all[kde])
                matrice_uni_LoQ.append(pd.concat([matrice_all[kde].iloc[:,0:2], matrice_all[kde].iloc[:,2:4]*koef], axis=1))  #matrice_uni_LoQ[0].iloc[1414]
                pocet_tranzice = matrice_uni_LoQ[i].iloc[:,2] - W_all[i].iloc[:,6]
                pocet_transverze = matrice_uni_LoQ[i].iloc[:,3] - W_all[i].iloc[:,7]

                tranzice_kde = np.where(pocet_tranzice< 0)[0]
                transverze_kde = np.where(pocet_transverze< 0)[0]
                rozmer = max(transverze_kde.size, tranzice_kde.size)
                if rozmer != 0:
                    M = np.zeros((rozmer, 11))
                    M[:tranzice_kde.size, :5] = np.concatenate((
                        matrice_uni_LoQ[i].iloc[tranzice_kde, [0,2]],
                        W_all[i].iloc[tranzice_kde, [6]],
                        matrice_uni_LoQ[i].iloc[tranzice_kde, [2]] / 30,
                        W_all[i].iloc[tranzice_kde, [6]] / 30
                    ), axis=1)
                    M[:transverze_kde.size, 6:11] = np.concatenate((
                        matrice_uni_LoQ[i].iloc[transverze_kde, [0, 3]],
                        W_all[i].iloc[transverze_kde, [7]],
                        matrice_uni_LoQ[i].iloc[transverze_kde, [3]] / 30,
                        W_all[i].iloc[transverze_kde, [7]] / 30
                    ), axis=1)
                    M[:, 5] = 0
                else:
                    M = np.zeros((rozmer, 11))
                popis = [['tranzice', 'cetnosti', '', '      %      ', '', '', 'transverze', 'cetnosti', '', '      %      ', ''],
                ['', 'app.matrice', 'sample', 'app.matrice', 'sample', '', '', 'app.matrice', 'sample', 'app.matrice', 'sample']]
                sizeM = M.shape[0]
                rozsah = 'T6:AD' + str(sizeM + 5)              
                logger.info('Sheet Completed: ' + str(i+1)+'/'+str(len(kteri)) )
                if sizeM != 0:
                    workbook = openpyxl.load_workbook(vystup)
                    sheet_names = workbook.sheetnames
                    if workbook.active != sheet_names[i]:
                        workbook.active = workbook[sheet_names[i]]
                    # select the active sheet
                    active_sheet = workbook.active

                    # update the range with the values from M
                    if sizeM > 0:
                        active_sheet_range = active_sheet[rozsah]
                        for row in active_sheet_range:
                            for cell in row:
                                cell.value = M[cell.row - 6][cell.column - 20]

                    # update the range with the values from popis
                    popis_range = active_sheet['T4:AD5']
                    for row in popis_range:
                        for cell in row:
                            cell.value = popis[cell.row - 4][cell.column - 20]

                    # update the range with the values from pat
                    aa_range = active_sheet['R6:R' + str(s + 5)]
                    for row in aa_range:
                        for cell in row:
                            cell.value = pat[cell.row - 6]

                    # update the range with the values from mut
                    crm_range = active_sheet['S6:S' + str(s + 5)]
                    for row in crm_range:
                        for cell in row:
                            cell.value = mut[cell.row-6]

                    # update the CRM label
                    crm_label_range = active_sheet['S5:S5']
                    crm_label_range[0][0].value = 'CRM'

                    # update the range with the values from meze_popis
                    emeze_range = active_sheet['H1:N2']
                    for row in emeze_range:
                        for cell in row:
                            cell.value = meze_popis[i].iloc[cell.row - 1][cell.column - 8]

                    # save the changes
                    workbook.save(vystup)
                else:
                    workbook = openpyxl.load_workbook(vystup)
                    sheet_names = workbook.sheetnames
                    if workbook.active != sheet_names[i]:
                        workbook.active = workbook[sheet_names[i]]
                    # select the active sheet
                    active_sheet = workbook.active

                    crm_label_range = active_sheet['S5:S5']
                    crm_label_range[0][0].value = 'CRM'

                    emeze_range = active_sheet['H1:N2']
                    for row in emeze_range:
                        for cell in row:
                            cell.value = meze_popis[i].iloc[cell.row - 1][cell.column - 8]

                    # save the changes
                    workbook.save(vystup)
            else:
                workbook = openpyxl.load_workbook(vystup)
                sheet_names = workbook.sheetnames
                if workbook.active != sheet_names[i]:
                    workbook.active = workbook[sheet_names[i]]
                active_sheet = workbook.active
                epopis = active_sheet['T4:AD5']
                popis = [['tranzice', 'cetnosti', '', '      %      ', '', '', 'transverze', 'cetnosti', '', '      %      ', ''],
                ['', 'app.matrice', 'sample', 'app.matrice', 'sample', '', '', 'app.matrice', 'sample', 'app.matrice', 'sample']]
                emeze = meze_popis[i]

                popis_range = active_sheet['T4:AD5']
                for row in popis_range:
                    for cell in row:
                        cell.value = popis[cell.row - 4][cell.column - 20]
                emeze_range = active_sheet['H1:N2']
                for row in emeze_range:
                    for cell in row:
                        cell.value = meze_popis[i].iloc[cell.row - 1][cell.column - 8]
                workbook.save(vystup)
            if limits != "0":
                workbook = openpyxl.load_workbook(vystup)
                sheet_names = workbook.sheetnames
                if workbook.active != sheet_names[i]:
                    workbook.active = workbook[sheet_names[i]]
                active_sheet = workbook.active
                active_sheet.conditional_formatting.add('M1', CellIsRule(operator='greaterThan', formula=['500'], fill=PatternFill(patternType='solid', bgColor='FF0000')))
                active_sheet.conditional_formatting.add('M2', CellIsRule(operator='lessThan', formula=['1200'], fill=PatternFill(patternType='solid', bgColor='FF0000')))
                            
                active_sheet.conditional_formatting.add('A6:S108', FormulaRule(formula=['=IF($P6="Synonymous", IFERROR(VLOOKUP($C6,$T:$T,1,FALSE), FALSE), FALSE)'], fill=PatternFill(patternType='solid', bgColor='00FF00')))
                active_sheet.conditional_formatting.add('A6:S108', FormulaRule(formula=['=IF($P6="Synonymous", IFERROR(VLOOKUP($C6,$Z:$Z,1,FALSE), FALSE), FALSE)'], fill=PatternFill(patternType='solid', bgColor='00FF00')))
                
                active_sheet.conditional_formatting.add('A6:S108', FormulaRule(formula=['=AND($H6>2, NOT(ISERROR(MATCH($C6,$T:$T,0))))'], fill=PatternFill(patternType='solid', bgColor='FC9804')))
                active_sheet.conditional_formatting.add('A6:S108', FormulaRule(formula=['=AND($H6>2, NOT(ISERROR(MATCH($C6,$Z:$Z,0))))'], fill=PatternFill(patternType='solid', bgColor='FC9804')))

                active_sheet.conditional_formatting.add('A6:S108', FormulaRule(formula=['=VLOOKUP($C6,$T:$T,1,FALSE)'], fill=PatternFill(patternType='solid', bgColor='FFFF66')))
                active_sheet.conditional_formatting.add('A6:S108', FormulaRule(formula=['=VLOOKUP($C6,$Z:$Z,1,FALSE)'], fill=PatternFill(patternType='solid', bgColor='FFFF66')))

                
                

                
                active_sheet['Q1'].fill = PatternFill(patternType='solid', fgColor='00FF00')
                active_sheet['Q1'].value = "Synonymous"
                active_sheet['Q2'].fill = PatternFill(patternType='solid', fgColor='FC9804')
                active_sheet['Q2'].value = "above 2%"
                active_sheet['Q3'].fill = PatternFill(patternType='solid', fgColor='FFFF66')
                active_sheet['Q3'].value = "below 2%"
                workbook.save(vystup)
    except Exception as e:
        logger.error(e)
        error_message = "Pravděpodobně jste zadali špatný soubor/y.\nZkuste to znovu"
        raise Exception(error_message)
#if xls_csv == 1:
#    report_part1_save_xlsx()
#    report_part2_save_xlsx()
#    report_part3_save_xlsx()
#else:  # cesta k folder of healthy Donors
#    pass